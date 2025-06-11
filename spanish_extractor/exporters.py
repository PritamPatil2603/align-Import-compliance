import json
import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import ExtractionSummary, ESNData, InvoiceData, LineItem, ExtractionConfidence

logger = logging.getLogger(__name__)

class SpanishInvoiceExporter:
    """Export Spanish invoice extraction results to multiple formats"""
    
    def __init__(self, export_dir: str):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        
        # Color schemes for Excel
        self.confidence_colors = {
            'HIGH': 'C6EFCE',      # Light green
            'MEDIUM': 'FFEB9C',    # Light yellow
            'LOW': 'FFC7CE',       # Light orange
            'ERROR': 'FF9999'      # Light red
        }
    
    def export_all_formats(self, summary: ExtractionSummary) -> Dict[str, str]:
        """Export to all formats and return file paths"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        export_paths = {}
        
        try:
            # Export JSON
            if hasattr(summary, 'to_json'):
                json_path = self.export_json(summary, timestamp)
                export_paths['json'] = json_path
                
            # Export CSV (line items)
            csv_path = self.export_csv(summary, timestamp)
            export_paths['csv'] = csv_path
            
            # Export Excel (multiple sheets)
            excel_path = self.export_excel(summary, timestamp)
            export_paths['excel'] = excel_path
            
            logger.info(f"✅ Exported to {len(export_paths)} formats")
            
        except Exception as e:
            logger.error(f"❌ Export error: {e}")
            
        return export_paths
    
    def export_json(self, summary: ExtractionSummary, timestamp: str) -> str:
        """Export complete data to JSON"""
        json_path = self.export_dir / f"spanish_invoices_{timestamp}.json"
        
        try:
            # Use the summary's built-in JSON export
            json_data = summary.to_dict()
            
            # Add export metadata
            json_data['export_metadata'] = {
                'export_timestamp': datetime.now().isoformat(),
                'format_version': '1.0',
                'exported_by': 'Spanish Invoice Extractor'
            }
            
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"📄 JSON exported: {json_path}")
            return str(json_path)
            
        except Exception as e:
            logger.error(f"❌ JSON export failed: {e}")
            return ""
    
    def export_csv(self, summary: ExtractionSummary, timestamp: str) -> str:
        """Export line items to CSV (flattened format)"""
        csv_path = self.export_dir / f"spanish_invoices_line_items_{timestamp}.csv"
        
        try:
            # Prepare flattened data (one row per line item)
            csv_data = []
            
            for esn, esn_data in summary.esn_data.items():
                for invoice_file, invoice in esn_data.invoices.items():
                    
                    for line_item in invoice.line_items:
                        csv_data.append({
                            'Entry_Summary_Number': esn,
                            'PDF_File_Name': invoice_file,
                            'Invoice_Date': invoice.invoice_date,
                            'Supplier': invoice.supplier,
                            'Invoice_Number': invoice.invoice_number,
                            'Line_Number': line_item.line_number,
                            'SKU_Client_Reference': line_item.sku_client_reference,
                            'Material_Description': line_item.material_description,
                            'Customs_Quantity': line_item.customs_quantity,
                            'Customs_Unit': line_item.customs_unit,
                            'Tariff_Fraction': line_item.tariff_fraction,
                            'Unit_Value_USD': line_item.unit_value_usd,
                            'Total_Value_USD': float(line_item.total_value_usd),
                            'Extraction_Confidence': invoice.extraction_confidence.value,
                            'Processing_Time_Seconds': invoice.processing_time_seconds,
                            'Extraction_Notes': invoice.extraction_notes
                        })
            
            # Write CSV
            if csv_data:
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=csv_data[0].keys())
                    writer.writeheader()
                    writer.writerows(csv_data)
                
                logger.info(f"📊 CSV exported: {csv_path} ({len(csv_data)} line items)")
            else:
                logger.warning("No data to export to CSV")
            
            return str(csv_path)
            
        except Exception as e:
            logger.error(f"❌ CSV export failed: {e}")
            return ""
    
    def export_excel(self, summary: ExtractionSummary, timestamp: str) -> str:
        """Export to Excel with multiple formatted sheets"""
        excel_path = self.export_dir / f"spanish_invoices_{timestamp}.xlsx"
        
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                
                # Sheet 1: Line Items (detailed data)
                self._create_line_items_sheet(summary, writer)
                
                # Sheet 2: ESN Summary
                self._create_esn_summary_sheet(summary, writer)
                
                # Sheet 3: Processing Summary
                self._create_processing_summary_sheet(summary, writer)
                
                # Sheet 4: Audit Trail
                self._create_audit_trail_sheet(summary, writer)
            
            # Apply formatting
            self._apply_excel_formatting(excel_path)
            
            logger.info(f"📈 Excel exported: {excel_path}")
            return str(excel_path)
            
        except Exception as e:
            logger.error(f"❌ Excel export failed: {e}")
            return ""
    
    def _create_line_items_sheet(self, summary: ExtractionSummary, writer):
        """Create detailed line items sheet"""
        
        line_items_data = []
        
        for esn, esn_data in summary.esn_data.items():
            for invoice_file, invoice in esn_data.invoices.items():
                
                for line_item in invoice.line_items:
                    line_items_data.append({
                        'Entry Summary Number': esn,
                        'PDF File Name': invoice_file,
                        'Invoice Date': invoice.invoice_date,
                        'Supplier': invoice.supplier,
                        'Invoice Number': invoice.invoice_number,
                        'Line Number': line_item.line_number,
                        'SKU/Client Reference': line_item.sku_client_reference,
                        'Material Description': line_item.material_description,
                        'Customs Quantity': line_item.customs_quantity,
                        'Customs Unit': line_item.customs_unit,
                        'Tariff Fraction': line_item.tariff_fraction,
                        'Unit Value USD': line_item.unit_value_usd,
                        'Total Value USD': float(line_item.total_value_usd),
                        'Extraction Confidence': invoice.extraction_confidence.value,
                        'Processing Time': f"{invoice.processing_time_seconds:.1f}s",
                        'Notes': invoice.extraction_notes
                    })
        
        if line_items_data:
            df = pd.DataFrame(line_items_data)
            df.to_excel(writer, sheet_name='Invoice Line Items', index=False)
        else:
            # Create empty sheet with headers
            empty_df = pd.DataFrame(columns=['Entry Summary Number', 'PDF File Name', 'Invoice Date', 'Supplier'])
            empty_df.to_excel(writer, sheet_name='Invoice Line Items', index=False)
    
    def _create_esn_summary_sheet(self, summary: ExtractionSummary, writer):
        """Create ESN summary sheet"""
        
        esn_summary_data = []
        
        for esn, esn_data in summary.esn_data.items():
            esn_summary_data.append({
                'Entry Summary Number': esn,
                'Total Invoices': esn_data.total_invoices,
                'Total Line Items': esn_data.total_line_items,
                'Total Declared Value USD': float(esn_data.total_declared_value_usd),
                'Success Rate %': f"{esn_data.get_success_rate():.1f}%",
                'Processing Status': esn_data.processing_status,
                'Unique Suppliers': ', '.join(esn_data.get_unique_suppliers()),
                'Date Range': f"{esn_data.get_date_range()[0]} to {esn_data.get_date_range()[1]}" if esn_data.get_date_range()[0] else "N/A",
                'Confidence Distribution': self._format_confidence_distribution(esn_data.get_confidence_distribution())
            })
        
        if esn_summary_data:
            df = pd.DataFrame(esn_summary_data)
            df.to_excel(writer, sheet_name='ESN Summary', index=False)
    
    def _create_processing_summary_sheet(self, summary: ExtractionSummary, writer):
        """Create processing summary sheet"""
        
        processing_stats = summary.get_processing_stats()
        confidence_summary = summary.get_confidence_summary()
        
        summary_data = [
            ['Metric', 'Value', 'Details'],
            ['Processing Date', summary.timestamp, 'Full extraction run'],
            ['Total ESNs Processed', summary.total_esns, 'Entry Summary Numbers found'],
            ['Total PDFs Processed', summary.total_invoices, 'Commercial invoices only'],
            ['Total Line Items Extracted', summary.total_line_items, 'Individual SKU entries'],
            ['Overall Success Rate', f"{summary.success_rate_percentage:.1f}%", 'Line item extraction success'],
            ['Total Processing Time', f"{summary.processing_time_seconds:.1f} seconds", 'Including downloads'],
            ['Total Declared Value', f"${summary.get_total_declared_value():,.2f}", 'Sum of all line items'],
            ['Average Items per Invoice', f"{processing_stats.get('avg_processing_time', 0):.1f}", 'SKUs per invoice'],
            ['Average Processing Time', f"{processing_stats.get('avg_processing_time', 0):.1f}s", 'Per PDF'],
            ['Confidence Distribution', self._format_confidence_distribution(confidence_summary.get('overall', {})), 'Extraction quality']
        ]
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Processing Summary', index=False, header=False)
    
    def _create_audit_trail_sheet(self, summary: ExtractionSummary, writer):
        """Create audit trail sheet for compliance"""
        
        audit_data = []
        
        for esn, esn_data in summary.esn_data.items():
            for invoice_file, invoice in esn_data.invoices.items():
                
                for line_item in invoice.line_items:
                    audit_data.append({
                        'ESN': esn,
                        'PDF File': invoice_file,
                        'Line Item': line_item.line_number,
                        'SKU': line_item.sku_client_reference,
                        'Declared Quantity': line_item.customs_quantity,
                        'Declared Value': float(line_item.total_value_usd),
                        'Unit Price': line_item.unit_value_usd,
                        'Tariff Code': line_item.tariff_fraction,
                        'Extraction Timestamp': summary.timestamp,
                        'Validator': 'AI_REGEX',
                        'Status': 'VERIFIED' if invoice.extraction_confidence in [ExtractionConfidence.HIGH, ExtractionConfidence.MEDIUM] else 'REVIEW_NEEDED'
                    })
        
        if audit_data:
            df = pd.DataFrame(audit_data)
            df.to_excel(writer, sheet_name='Audit Trail', index=False)
    
    def _apply_excel_formatting(self, excel_path: str):
        """Apply professional formatting to Excel file"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = load_workbook(excel_path)
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Header formatting
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin")
                        )
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Apply confidence color coding to Line Items sheet
                if sheet_name == 'Invoice Line Items':
                    self._apply_confidence_colors(ws)
                
                # Freeze top row
                if ws.max_row > 1:
                    ws.freeze_panes = 'A2'
            
            wb.save(excel_path)
            logger.debug("✅ Excel formatting applied")
            
        except Exception as e:
            logger.warning(f"⚠️ Excel formatting failed: {e}")
    
    def _apply_confidence_colors(self, worksheet):
        """Apply color coding based on confidence levels"""
        try:
            # Find confidence column
            confidence_col = None
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value and 'confidence' in str(cell_value).lower():
                    confidence_col = col
                    break
            
            if confidence_col:
                for row in range(2, worksheet.max_row + 1):
                    confidence_value = worksheet.cell(row=row, column=confidence_col).value
                    
                    if confidence_value in self.confidence_colors:
                        fill_color = self.confidence_colors[confidence_value]
                        
                        # Apply color to entire row
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
        except Exception as e:
            logger.warning(f"⚠️ Confidence color coding failed: {e}")
    
    def _format_confidence_distribution(self, distribution: Dict[str, int]) -> str:
        """Format confidence distribution for display"""
        if not distribution:
            return "N/A"
        
        total = sum(distribution.values())
        if total == 0:
            return "No data"
        
        formatted_parts = []
        for confidence, count in distribution.items():
            if count > 0:
                percentage = (count / total) * 100
                formatted_parts.append(f"{confidence}: {percentage:.1f}%")
        
        return ", ".join(formatted_parts)
    
    def export_summary_report(self, summary: ExtractionSummary, timestamp: str) -> str:
        """Export a human-readable summary report"""
        report_path = self.export_dir / f"spanish_invoices_summary_{timestamp}.txt"
        
        try:
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("SPANISH INVOICE EXTRACTION SUMMARY REPORT\n")
                f.write("=" * 50 + "\n\n")
                
                f.write(f"Extraction Date: {summary.timestamp}\n")
                f.write(f"Total ESNs Processed: {summary.total_esns}\n")
                f.write(f"Total Invoices: {summary.total_invoices}\n")
                f.write(f"Total Line Items: {summary.total_line_items}\n")
                f.write(f"Success Rate: {summary.success_rate_percentage:.1f}%\n")
                f.write(f"Total Declared Value: ${summary.get_total_declared_value():,.2f}\n\n")
                
                # ESN breakdown
                f.write("ESN BREAKDOWN:\n")
                f.write("-" * 30 + "\n")
                
                for esn, esn_data in summary.esn_data.items():
                    f.write(f"\n{esn}:\n")
                    f.write(f"  Invoices: {esn_data.total_invoices}\n")
                    f.write(f"  Line Items: {esn_data.total_line_items}\n")
                    f.write(f"  Total Value: ${esn_data.total_declared_value_usd:,.2f}\n")
                    f.write(f"  Status: {esn_data.processing_status}\n")
                
                # Processing stats
                processing_stats = summary.get_processing_stats()
                f.write("\nPROCESSING STATISTICS:\n")
                f.write("-" * 30 + "\n")
                f.write(f"Average Processing Time: {processing_stats.get('avg_processing_time', 0):.1f}s\n")
                f.write(f"Total Processing Time: {processing_stats.get('total_processing_time', 0):.1f}s\n")
                
                # Confidence summary
                confidence_summary = summary.get_confidence_summary()
                f.write("\nCONFIDENCE SUMMARY:\n")
                f.write("-" * 30 + "\n")
                overall_dist = confidence_summary.get('overall', {})
                for conf, count in overall_dist.items():
                    f.write(f"{conf}: {count} invoices\n")
            
            logger.info(f"📋 Summary report exported: {report_path}")
            return str(report_path)
            
        except Exception as e:
            logger.error(f"❌ Summary report export failed: {e}")
            return ""